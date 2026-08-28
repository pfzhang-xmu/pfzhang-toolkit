#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Data-to-Paper: 从实验数据自动生成统计、图表与 Results 草稿。

CLI:
  python data2paper.py stats <file> [--group col] [--values a,b,c]
  python data2paper.py charts <file> [--group col] [--values a,b,c] [--out dir]
  python data2paper.py fill <project_dir> [--values a,b,c]
  python data2paper.py figures <project_dir> [--mode auto|all] [--ids 图1,图3] [--json]
"""
import argparse
import csv
import io
import math
import re
import statistics
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

try:
    import pandas as pd
    HAS_PANDAS = True
except Exception:
    HAS_PANDAS = False

try:
    from scipy import stats as sp_stats
    HAS_SCIPY = True
except Exception:
    HAS_SCIPY = False


def read_table(path):
    """读取 CSV / XLSX，返回 (columns, rows[dict])。"""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = [dict(r) for r in reader]
        columns = list(rows[0].keys()) if rows else []
        return columns, rows
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        wb.close()
        if not data:
            return [], []
        header = [str(c).strip() if c is not None else "" for c in data[0]]
        rows = []
        for r in data[1:]:
            if any(v is not None and str(v).strip() != "" for v in r):
                rows.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
        columns = header
        return columns, rows
    raise ValueError(f"不支持的文件类型: {suffix}（支持 .csv / .xlsx）")


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def numeric_columns(columns, rows):
    out = []
    for c in columns:
        vals = [_num(r.get(c)) for r in rows]
        if any(v is not None for v in vals):
            out.append(c)
    return out


def describe(rows, col):
    vals = [_num(r.get(col)) for r in rows]
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    return {
        "count": len(vals),
        "mean": round(statistics.mean(vals), 4),
        "sd": round(statistics.stdev(vals), 4) if len(vals) > 1 else 0.0,
        "min": round(min(vals), 4),
        "max": round(max(vals), 4),
        "median": round(statistics.median(vals), 4),
    }


def group_values(rows, group_col, value_col):
    groups = {}
    for r in rows:
        g = str(r.get(group_col, "")).strip() or "NA"
        v = _num(r.get(value_col))
        if v is None:
            continue
        groups.setdefault(g, []).append(v)
    return groups


def group_stats(rows, group_col, value_col):
    groups = group_values(rows, group_col, value_col)
    out = {}
    for g, vals in groups.items():
        out[g] = {
            "n": len(vals),
            "mean": round(statistics.mean(vals), 4),
            "sd": round(statistics.stdev(vals), 4) if len(vals) > 1 else 0.0,
        }
    return out


def significance(groups):
    """对两组做 Welch t 检验；对 >=3 组做单因素 ANOVA（若 scipy 可用）。

    小样本（任一组 n<3）时不输出 p 值，只给提示，避免误导性统计结论；
    组内无变异（所有值相同）时返回 none 提示，避免 F=nan 脏输出；
    正常检验附带效应量（Cohen's d / η²）。
    """
    if not HAS_SCIPY:
        return None
    items = list(groups.items())
    sizes = [len(v) for _, v in items]
    if any(n < 3 for n in sizes):
        return {"test": "none", "warning": f"样本量过小（各组 n={sizes}），不输出显著性检验，建议每组 n≥3", "p": None}
    if len(items) == 2:
        a = items[0][1]
        b = items[1][1]
        if all(x == a[0] for x in a) and all(x == b[0] for x in b) and a[0] == b[0]:
            return {"test": "none", "warning": "组间无差异（所有值相同），无法做显著性检验", "p": None}
        t, p = sp_stats.ttest_ind(a, b, equal_var=False)
        if p != p:  # nan 防护
            return {"test": "none", "warning": "组间无变异，无法做显著性检验", "p": None}
        na, nb = len(a), len(b)
        sa, sb = statistics.stdev(a), statistics.stdev(b)
        spool = math.sqrt(((na - 1) * sa ** 2 + (nb - 1) * sb ** 2) / (na + nb - 2)) if na + nb > 2 else 0.0
        d = (statistics.mean(a) - statistics.mean(b)) / spool if spool else 0.0
        return {"test": "Welch t-test", "statistic": round(t, 4), "p": float(p),
                "effect_size": {"cohens_d": round(d, 4)}}
    if len(items) >= 3:
        f, p = sp_stats.f_oneway(*[v for _, v in items])
        if p != p or f != f:  # nan 防护（组内无变异）
            return {"test": "none", "warning": "组内无变异（所有值相同），无法做显著性检验", "p": None}
        # η² = SS_between / SS_total
        allv = [x for _, v in items for x in v]
        grand = statistics.mean(allv)
        ss_between = sum(len(v) * (statistics.mean(v) - grand) ** 2 for _, v in items)
        ss_total = sum((x - grand) ** 2 for x in allv)
        eta2 = (ss_between / ss_total) if ss_total else 0.0
        return {"test": "one-way ANOVA", "statistic": round(f, 4), "p": float(p),
                "effect_size": {"eta_squared": round(eta2, 4)}}
    return None


def generate_markdown(path, group_col=None, value_cols=None):
    columns, rows = read_table(path)
    if not rows:
        return "数据为空"
    if value_cols is None:
        value_cols = numeric_columns(columns, rows)
    lines = [f"# 自动生成 Results（数据文件: {Path(path).name}）", ""]
    if group_col and group_col in columns:
        lines.append(f"分组变量: **{group_col}**")
        lines.append("")
        for vc in value_cols:
            if vc not in columns:
                continue
            gs = group_stats(rows, group_col, vc)
            raw = group_values(rows, group_col, vc)
            lines.append(f"## {vc}")
            lines.append("")
            lines.append("| 组 | n | 均值 | SD |")
            lines.append("|---|----|------|----|")
            for g, s in gs.items():
                lines.append(f"| {g} | {s['n']} | {s['mean']} | {s['sd']} |")
            sig = significance(raw)
            if sig:
                lines.append("")
                if sig.get("p") is None:
                    lines.append(f"⚠ {sig['warning']}")
                else:
                    line = f"统计检验: {sig['test']}, statistic = {sig['statistic']}, p = {sig['p']:.4g}"
                    if sig.get("effect_size"):
                        for k, v in sig["effect_size"].items():
                            line += f", {k} = {v}"
                    lines.append(line)
            lines.append("")
    else:
        lines.append("| 指标 | n | 均值 | SD | 中位数 | 最小 | 最大 |")
        lines.append("|------|---|------|----|--------|------|------|")
        for vc in value_cols:
            if vc not in columns:
                continue
            d = describe(rows, vc)
            if d:
                lines.append(f"| {vc} | {d['count']} | {d['mean']} | {d['sd']} | {d['median']} | {d['min']} | {d['max']} |")
        lines.append("")
    lines.append("> 注: 此文件由 data2paper.py 自动生成，需人工核对实验设计与统计假设。")
    return "\n".join(lines)


def generate_charts(path, group_col=None, value_cols=None, out_dir=None):
    from web.charts import generate_chart
    columns, rows = read_table(path)
    if value_cols is None:
        value_cols = numeric_columns(columns, rows)
    content = Path(path).read_text(encoding="utf-8-sig", errors="replace")
    results = []
    for vc in value_cols[:6]:
        chart_type = "bar" if group_col and group_col in columns else "line"
        x = group_col if group_col and group_col in columns else columns[0]
        res = generate_chart(Path(path).name, content, chart_type, x, vc, vc, out_dir)
        results.append(res)
    return results


def parse_figures_plan(project):
    """解析 framework/figures.md 的图表规划表，返回 dict 列表。"""
    plan_file = Path(project) / "framework" / "figures.md"
    if not plan_file.exists():
        return []
    rows = []
    header = None
    for raw_line in plan_file.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.lstrip("\ufeff")
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if header is None:
            header = cells
            continue
        if len(cells) < 2:
            continue
        if all(set(c) <= set("-: ") for c in cells if c):
            continue
        rec = dict(zip(header, cells))
        if rec.get("编号", "").startswith(("图", "表")):
            rows.append(rec)
    return rows


def _clean_source(src):
    if not src:
        return ""
    s = src.strip().replace("data/", "").replace("\\", "/")
    for ext in (".csv", ".xlsx", ".xls"):
        if s.lower().endswith(ext):
            s = s[: -len(ext)]
    return s.strip()


def find_data_file(data_dir, source):
    """按 figures.md 里的‘数据来源’模糊匹配 data/ 下的 CSV/XLSX。"""
    if not source:
        return None
    key = _clean_source(source)
    if not key:
        return None
    candidates = [p for p in sorted(data_dir.iterdir()) if p.suffix.lower() in (".csv", ".xlsx")]
    for p in candidates:
        if p.stem == key or key in p.stem or p.stem in key:
            return p
    tokens = [t for t in re.split(r"[/\s,，;；、]+", key) if len(t) >= 2]
    for p in candidates:
        for t in tokens:
            if t in p.stem:
                return p
    return None


def chart_type_from_requirement(req):
    if not req:
        return "bar"
    r = req.lower()
    if "折线" in r or "line" in r:
        return "line"
    if "散点" in r or "scatter" in r:
        return "scatter"
    if "饼" in r or "pie" in r:
        return "pie"
    if "箱线" in r or "box" in r:
        return "box"
    if "小提琴" in r or "violin" in r:
        return "violin"
    if "直方" in r or "hist" in r or "分布" in r:
        return "hist"
    if "面积" in r or "area" in r or "填充" in r:
        return "area"
    if "热力" in r or "heatmap" in r or "相关" in r:
        return "heatmap"
    return "bar"


def _pick_value_column(content, numeric_cols):
    if not numeric_cols:
        return None
    keywords = ["产量", "再生率", "活性", "含量", "浓度", "效价", "稳定性", "biomass",
                "yield", "value", "growth", "titer", "活性成分", "多糖", "腺苷", "虫草素"]
    low_content = (content or "").lower()
    for c in numeric_cols:
        cl = c.lower()
        if any(k in cl for k in keywords):
            return c
    for c in numeric_cols:
        if c.lower() in low_content:
            return c
    return numeric_cols[-1]


def _table_markdown(columns, rows):
    """根据数据生成一个简洁的统计表 Markdown。"""
    numeric = numeric_columns(columns, rows)
    if not numeric:
        return None
    group_col = next((c for c in columns if c not in numeric), None)
    if group_col:
        lines = ["| {} | 指标 | n | 均值 | SD |".format(group_col), "|---|---|---|---|---|"]
        groups = {}
        for r in rows:
            g = str(r.get(group_col, "")).strip() or "NA"
            groups.setdefault(g, []).append(r)
        for g in sorted(groups):
            for vc in numeric:
                vals = [_num(r.get(vc)) for r in groups[g]]
                vals = [v for v in vals if v is not None]
                if vals:
                    mean = round(statistics.mean(vals), 4)
                    sd = round(statistics.stdev(vals), 4) if len(vals) > 1 else 0.0
                    lines.append(f"| {g} | {vc} | {len(vals)} | {mean} | {sd} |")
    else:
        lines = ["| 指标 | n | 均值 | SD | 中位数 | 最小 | 最大 |", "|---|---|---|---|---|---|---|"]
        for vc in numeric:
            vals = [_num(r.get(vc)) for r in rows]
            vals = [v for v in vals if v is not None]
            if vals:
                lines.append(
                    f"| {vc} | {len(vals)} | {round(statistics.mean(vals), 4)} | "
                    f"{round(statistics.stdev(vals), 4) if len(vals) > 1 else 0} | "
                    f"{round(statistics.median(vals), 4)} | {round(min(vals), 4)} | {round(max(vals), 4)} |"
                )
    return "\n".join(lines)


def _anova_table(columns, rows, group_col, numeric):
    """按分组变量对每个数值列做单因素 ANOVA，输出 F / p / η² 表。"""
    lines = [f"| {group_col} 组间 | 指标 | F | p | η² |", "|---|---|---|---|---|"]
    for vc in numeric:
        groups = group_values(rows, group_col, vc)
        sig = significance(groups)
        if sig and sig.get("p") is not None:
            es = sig.get("effect_size", {}) or {}
            eta = es.get("eta_squared", "")
            lines.append(f"| {group_col} | {vc} | {sig['statistic']} | {sig['p']:.4g} | {eta} |")
        else:
            note = (sig or {}).get("warning", "scipy 不可用")
            lines.append(f"| {group_col} | {vc} | - | - | {note} |")
    lines.append("")
    lines.append("> 注: 自动单因素 ANOVA(组间);多因素交互需 statsmodels 或人工补充。")
    return "\n".join(lines)


def generate_planned_tables(project, data_files=None):
    """按 framework/figures.md 的‘表’规划，为每张表生成统计表 Markdown。

    表类型按规划行的『内容/期刊规范要求』识别：
    - 含 ANOVA/方差分析/显著性检验 → 生成单因素 ANOVA 表(F/p/η²)
    - 其余 → 描述统计表(均值/SD)
    同一数据源+同一分组只生成一次(防止 表1/表3 规划重复时插入两张相同表)。
    """
    project = Path(project)
    data_dir = project / "data"
    if data_files is None:
        data_files = [p for p in sorted(data_dir.iterdir()) if p.suffix.lower() in (".csv", ".xlsx")]
    plan = parse_figures_plan(project)
    tables_info = []
    seen_sig = set()
    for rec in plan:
        tab = rec.get("编号", "").strip()
        ftype = (rec.get("类型") or rec.get("类型(图/表)") or "").strip()
        if not tab.startswith("表") or "表" not in ftype:
            continue
        content = (rec.get("内容", "") or "") + " " + (rec.get("期刊规范要求", "") or "")
        data_file = find_data_file(data_dir, rec.get("数据来源", ""))
        if data_file is None:
            continue
        try:
            columns, rows = read_table(data_file)
        except Exception:
            continue
        numeric = numeric_columns(columns, rows)
        if not numeric:
            continue
        group_col = next((c for c in columns if c not in numeric), None)
        sig_key = (data_file.name, group_col)
        is_anova = bool(re.search(r"ANOVA|方差分析|显著性检验|F检验|f\s*test", content, re.I)) and HAS_SCIPY
        if is_anova and group_col:
            md = _anova_table(columns, rows, group_col, numeric)
            if md:
                tables_info.append({
                    "table": tab, "section": rec.get("对应章节", ""),
                    "markdown": md, "source": data_file.name, "kind": "anova",
                })
                seen_sig.add(sig_key)
            continue
        if sig_key in seen_sig:
            # 同源同分组已生成过描述统计表 → 跳过重复,避免 表N 与已插入表内容相同
            tables_info.append({
                "table": tab, "section": rec.get("对应章节", ""),
                "markdown": None, "source": data_file.name, "kind": "dup",
            })
            continue
        md = _table_markdown(columns, rows)
        if not md:
            continue
        tables_info.append({
            "table": tab, "section": rec.get("对应章节", ""),
            "markdown": md, "source": data_file.name, "kind": "desc",
        })
        seen_sig.add(sig_key)
    return tables_info


def _ref_en(ref):
    """图2 → Figure 2；表2 → Table 2。"""
    if ref.startswith("表"):
        return "Table " + ref[1:]
    return "Figure " + ref[1:]


def _strip_old_auto(text, ref):
    """移除正文中该图表旧的自动生成块，使 fill 重跑能刷新为最新版本。

    优先按边界标记删除（<!-- AUTO-TABLE:表3 --> ... <!-- /AUTO-TABLE -->）；
    兼容无标记的旧格式：表块 = 「**表N（自动生成...）**」标题 + 后续表格/注释行，
    以及历史上遗落的无标题表格残留（孤立的连续表格行块 ≥3 行 + 可选注释行）。
    """
    if ref.startswith("表"):
        # 带标记块
        text = re.sub(
            r"<!--\s*AUTO-TABLE:" + re.escape(ref) + r"\s*-->\s*\n?[\s\S]*?<!--\s*/AUTO-TABLE\s*-->\s*\n?",
            "", text)
        # 无标记旧格式: 带标题块
        text = re.sub(
            r"\*\*" + re.escape(ref) + r"（自动生成[^）]*）\*\*\n(?:\s*\n?)*(?:(?:\|.*\n?)|(?:>.*\n?))*",
            "", text)
        # 无标题残留: 孤立连续表格行块(≥3 行)+ 可选注释行(允许其间空行),前后为空行/段落边界
        text = re.sub(
            r"\n\n(?:\|.*\n){3,}(?:\s*\n)*(?:>.*\n)*(?=\n|\Z)", "\n\n", text)
        # 历史残留: 孤立的 data2paper 注释行(含「自动」字样)散落正文
        text = re.sub(
            r"\n\n>\s*[^\n]*(?:自动|data2paper)[^\n]*\n(?=\n|\Z)", "\n\n", text)
    else:
        # 带标记块
        text = re.sub(
            r"<!--\s*AUTO-FIGURE:" + re.escape(ref) + r"\s*-->\s*\n?[\s\S]*?<!--\s*/AUTO-FIGURE\s*-->\s*\n?",
            "", text)
        # 无标记旧格式
        text = re.sub(
            r"\[AUTO\]\s*" + re.escape(ref) + r"[^\n]*\n?!\[[^\]]*\]\([^\)]*\)\s*\n?",
            "", text)
    return text


def insert_planned_into_manuscript(project, charts_info, tables_info):
    """按 figures.md 规划顺序，一次性把图和表插入 manuscript/main.md。

    修复：同一行 [TBD: ... Figure 2 and Table 2 ...] 命中多个图表时，
    按规划顺序把该行一次性替换为所有图表的拼接块（先图后表、编号升序），
    避免先插入者吞掉整行、后续图表退位导致顺序颠倒。
    """
    main_md = Path(project) / "manuscript" / "main.md"
    if not main_md.exists() or (not charts_info and not tables_info):
        return (0, 0)
    plan = parse_figures_plan(project)
    order = {}
    for i, rec in enumerate(plan):
        order[rec.get("编号", "").strip()] = i
    try:
        from web.imagegen_bridge import DISCLOSURE
    except Exception:
        DISCLOSURE = "Illustration generated with an AI image model."
    blocks = []
    for c in charts_info:
        ref = c.get("figure", "")
        if not ref:
            continue
        try:
            caption = (c.get("content") or "").strip()
            if c.get("text_md"):
                # text 路线（含 imagegen 降级）：结构化文字图注块，无图片引用
                body = c["text_md"]
            else:
                rel = c.get("rel", "")
                if not rel or not Path(c.get("file", "")).exists():
                    continue
                if c.get("route") == "imagegen":
                    # imagegen 路线：![图N 图注](rel) + 图注行附 DISCLOSURE 披露句
                    alt = ref + (" " + caption if caption else "")
                    cap = ("%s %s. %s" % (ref, caption, DISCLOSURE)) if caption else ("%s. %s" % (ref, DISCLOSURE))
                    body = "![%s](%s)\n*%s*" % (alt, rel, cap)
                else:
                    body = "[AUTO] %s 已由 data2paper 自动生成并插入（%s）。\n![%s](%s)" % (ref, rel, ref, rel)
        except Exception:
            continue
        blocks.append({
            "ref": ref, "kind": "figure",
            "order": order.get(ref, 10 ** 9),
            "section": c.get("section", ""),
            "text": "\n<!-- AUTO-FIGURE:{} -->\n{}\n<!-- /AUTO-FIGURE -->\n".format(ref, body),
        })
    for t in tables_info:
        ref = t.get("table", "")
        md = t.get("markdown", "")
        if not ref or not md:
            continue
        kind = t.get("kind", "desc")
        tag = "（自动生成 ANOVA 表）" if kind == "anova" else "（自动生成）"
        blocks.append({
            "ref": ref, "kind": "table",
            "order": order.get(ref, 10 ** 9),
            "section": t.get("section", ""),
            "text": "\n<!-- AUTO-TABLE:{} -->\n**{}{}**\n\n{}\n<!-- /AUTO-TABLE -->\n".format(ref, ref, tag, md),
        })
    blocks.sort(key=lambda b: b["order"])
    text = main_md.read_text(encoding="utf-8", errors="replace")
    backup = main_md.with_name("main.md.bak")
    if not backup.exists():
        backup.write_text(text, encoding="utf-8")
    inserted = {"figure": 0, "table": 0}

    def _hit(line, blk):
        # 匹配「图2/Figure 2/表2/Table 2」，且编号后不是数字（避免图2 误命中图20）
        return re.search(
            re.escape(blk["ref"]) + r"(?!\d)|" + re.escape(_ref_en(blk["ref"])) + r"(?!\d)",
            line,
        )

    # 1) 按 TBD 行分组替换：一行命中多个图表时按规划顺序拼接，一次替换
    def _replace_tbd(match):
        line = match.group(0)
        hits = [b for b in blocks if not b.get("_done") and _hit(line, b)]
        if not hits:
            return line
        for b in hits:
            b["_done"] = True
            inserted[b["kind"]] += 1
        return "".join(b["text"] for b in hits)

    text = re.sub(r"\[TBD:[^\n]*\]", _replace_tbd, text)

    # 2) 未命中 TBD 的图表：先剥离旧的自动生成块（刷新版本），再按对应章节标题插入
    for blk in blocks:
        if blk.get("_done"):
            continue
        ref = blk["ref"]
        text = _strip_old_auto(text, ref)
        sec_key = None
        if blk.get("section"):
            msec = re.search(r"(\d+(?:\.\d+)?)", str(blk["section"]))
            if msec:
                sec_key = msec.group(1)
        placed = False
        if sec_key:
            heading = re.compile(r"^(#{2,4}\s*" + re.escape(sec_key) + r"\b.*)$", re.MULTILINE)
            hm = heading.search(text)
            if hm:
                if ref not in text[hm.end():hm.end() + 3000]:
                    text = text[:hm.end()] + "\n" + blk["text"] + text[hm.end():]
                    inserted[blk["kind"]] += 1
                    placed = True
        # 3) 兜底：插入 Results 区段末尾
        if not placed and "## 4. Results" in text and ref not in text:
            idx = text.find("## 5.")
            if idx == -1:
                idx = len(text)
            text = text[:idx] + "\n" + blk["text"] + "\n" + text[idx:]
            inserted[blk["kind"]] += 1
    main_md.write_text(text, encoding="utf-8")
    return inserted["figure"], inserted["table"]


def insert_tables_into_manuscript(project, tables_info):
    """把自动生成的统计表插入 manuscript/main.md（委托统一插入函数）。"""
    return insert_planned_into_manuscript(project, [], tables_info)[1]


def _prefer_png_over_svg(res, out_dir, stem):
    """data 路线选图：data/charts 下同 stem 存在 .png 时忽略 .svg（png 优先）。"""
    if (res.get("format") or "").lower() != "svg":
        return res
    from web.charts import _chart_stem_match
    try:
        pngs = [p for p in Path(out_dir).iterdir()
                if p.suffix.lower() == ".png" and _chart_stem_match(p.stem, stem)
                and p.stat().st_size > 0]
    except OSError:
        return res
    if not pngs:
        return res
    png = max(pngs, key=lambda p: p.stat().st_mtime)
    return {**res, "format": "png", "file": str(png), "rel": "data/charts/" + png.name}


def _figure_text_entry(rec, fig, project, pending_prefix=""):
    """text 路线（含 imagegen 降级）条目：结构化文字图注 + 块尾『待补图』标注。"""
    from web.charts import figure_caption_block
    note = "（待补图：{}自动生成可运行 `python data2paper.py figures {}`）".format(
        pending_prefix, Path(project).name)
    return {
        "figure": fig,
        "section": rec.get("对应章节", ""),
        "route": "text",
        "content": (rec.get("内容") or "").strip(),
        "text_md": figure_caption_block(rec).rstrip() + "\n\n" + note,
    }


def _figure_imagegen_entry(rec, fig, project):
    """imagegen 路线条目：复用 data/figures/fig_imagegen_<sha1(关键视觉)[:8]>.png；无产物降级 text 并标注待补图。"""
    from web.imagegen_bridge import cache_key_for
    key_visual = (rec.get("关键视觉") or "").strip()
    if key_visual:
        from web.imagegen_bridge import find_cached_figure_image
        fpath = find_cached_figure_image(Path(project) / "data" / "figures",
                                         cache_key_for(key_visual))
        if fpath is not None:
            return {
                "figure": fig,
                "section": rec.get("对应章节", ""),
                "route": "imagegen",
                "content": (rec.get("内容") or "").strip(),
                "file": str(fpath),
                "rel": "data/figures/" + fpath.name,
            }
    return _figure_text_entry(rec, fig, project, pending_prefix="imagegen 产物缺失，已降级文字图注；")


def generate_planned_charts(project, data_files=None):
    """按 framework/figures.md 的图表规划，为每个‘图’生成一张图（三路线）。

    路线由 web.charts.judge_figure_route 判定：
    - data：保留既有自动绘图逻辑；data/charts 下同 stem 的 png 优先于 svg；
    - imagegen：复用 data/figures/fig_imagegen_<sha1(关键视觉)[:8]>.png 产物；
      产物缺失 → 降级 text 并标注『待补图』；
    - text：charts.figure_caption_block 结构化文字图注（块尾标注『待补图』）。
    单图失败 try/except 跳过，不影响其他图。

    返回 [{figure, section, route, file/rel 或 text_md, ...}, ...]
    """
    project = Path(project)
    data_dir = project / "data"
    if data_files is None:
        data_files = [p for p in sorted(data_dir.iterdir()) if p.suffix.lower() in (".csv", ".xlsx")]
    from web.charts import generate_chart, judge_figure_route
    plan = parse_figures_plan(project)
    out_dir = project / "data" / "charts"
    out_dir.mkdir(parents=True, exist_ok=True)
    charts_info = []
    for rec in plan:
        fig = rec.get("编号", "").strip()
        ftype = (rec.get("类型") or rec.get("类型(图/表)") or "").strip()
        if not fig.startswith("图") or "图" not in ftype:
            continue
        try:
            route, _reason = judge_figure_route(rec, project)
            if route == "imagegen":
                info = _figure_imagegen_entry(rec, fig, project)
            elif route == "text":
                info = _figure_text_entry(rec, fig, project)
            else:  # data 路线：既有自动绘图逻辑
                data_file = find_data_file(data_dir, rec.get("数据来源", ""))
                if data_file is None:
                    continue
                columns, rows = read_table(data_file)
                numeric = numeric_columns(columns, rows)
                if not numeric:
                    continue
                x = next((c for c in columns if c not in numeric), columns[0])
                y = _pick_value_column(rec.get("内容", ""), numeric)
                if y is None:
                    continue
                req = (rec.get("期刊规范要求", "") or "") + " " + (rec.get("内容", "") or "")
                chart_type = chart_type_from_requirement(req)
                content = data_file.read_text(encoding="utf-8-sig", errors="replace")
                res = generate_chart(data_file.name, content, chart_type, x, y,
                                     f"{fig}: {rec.get('内容', '').strip()}", out_dir)
                res = _prefer_png_over_svg(res, out_dir, data_file.stem)
                info = {
                    **res,
                    "figure": fig,
                    "section": rec.get("对应章节", ""),
                    "source": data_file.name,
                    "route": "data",
                    "content": (rec.get("内容") or "").strip(),
                }
        except Exception:
            continue
        charts_info.append(info)
    return charts_info


def insert_charts_into_manuscript(project, charts_info):
    """把自动生成的图插入 manuscript/main.md（委托统一插入函数）。"""
    return insert_planned_into_manuscript(project, charts_info, [])[0]


# ── [DATA:] 占位符正文级回填 ──────────────────────────────
# 约定（确定性、可自动填充）:
#   [DATA: <数据文件>:<列>:<统计量>]            统计量: n|count|mean|sd|mean_sd|min|max|median|range
#   [DATA: <数据文件>:<分组列>:<数值列>:anova]   输出各组 mean±sd + 显著性 p
# 例: [DATA: fusion:rate:mean_sd]  [DATA: fusion:group:rate:anova]
# 数据文件按 stem 匹配 project/data/ 下的 .csv/.xlsx；无法匹配的占位符保留原样并计入 unfilled。
DATA_PH_RE = re.compile(r"\[DATA:\s*([^\]]+?)\s*\]")


def _match_col(name, columns):
    if not name:
        return None
    n = name.strip().lower()
    for c in columns:
        if str(c).strip().lower() == n:
            return c
    for c in columns:
        if n in str(c).strip().lower() or str(c).strip().lower() in n:
            return c
    return None


def _fmt_num(x):
    if x is None:
        return "NA"
    if isinstance(x, float):
        return ("%g" % round(x, 4))
    return str(x)


def _fmt_p(p):
    if p is None:
        return "p=NA"
    if p < 0.001:
        return "p<0.001"
    return "p=%.3f" % p


def _resolve_data_spec(spec, by_stem):
    parts = [p.strip() for p in spec.split(":") if p.strip() != ""]
    if len(parts) < 3:
        return None
    stem = parts[0].lower()
    f = by_stem.get(stem)
    if f is None:
        for k, v in by_stem.items():
            if stem in k or k in stem:
                f = v
                break
    if f is None:
        return None
    try:
        columns, rows = read_table(str(f))
    except Exception:
        return None
    if not rows:
        return None
    # 分组 + 显著性: file:group:value:anova
    if len(parts) >= 4 and parts[3].lower() in ("anova", "sig", "significance", "显著"):
        gcol = _match_col(parts[1], columns)
        vcol = _match_col(parts[2], columns)
        if not gcol or not vcol:
            return None
        gs = group_stats(rows, gcol, vcol)
        if not gs:
            return None
        seg = [f"{g}: {_fmt_num(s['mean'])}±{_fmt_num(s['sd'])} (n={s['n']})" for g, s in gs.items()]
        groups = group_values(rows, gcol, vcol)
        sig = significance(groups)
        if sig and sig.get("p") is not None:
            seg.append(f"{sig.get('test','')}, {_fmt_p(sig.get('p'))}")
        elif sig and sig.get("warning"):
            seg.append(sig["warning"])
        return "；".join(seg)
    # 单列统计: file:col:stat
    col = _match_col(parts[1], columns)
    if col is None:
        return None
    stat = parts[2].lower()
    d = describe(rows, col)
    if d is None:
        return None
    if stat in ("n", "count"):
        return _fmt_num(d["count"])
    if stat == "mean":
        return _fmt_num(d["mean"])
    if stat == "sd":
        return _fmt_num(d["sd"])
    if stat in ("mean_sd", "mean±sd"):
        return f"{_fmt_num(d['mean'])}±{_fmt_num(d['sd'])}"
    if stat == "min":
        return _fmt_num(d["min"])
    if stat == "max":
        return _fmt_num(d["max"])
    if stat == "median":
        return _fmt_num(d["median"])
    if stat == "range":
        return f"{_fmt_num(d['min'])}–{_fmt_num(d['max'])}"
    return None


def fill_data_placeholders(project, data_files=None):
    """回填 manuscript/main.md 正文中的 [DATA: ...] 占位符为真实统计值。

    返回 {filled, unfilled, items[未匹配占位符], error?}。无占位符时 filled=unfilled=0。
    """
    main_md = Path(project) / "manuscript" / "main.md"
    if not main_md.exists():
        return {"filled": 0, "unfilled": 0, "items": [], "error": "缺少 manuscript/main.md"}
    if data_files is None:
        data_dir = Path(project) / "data"
        data_files = ([p for p in sorted(data_dir.iterdir())
                       if p.suffix.lower() in (".csv", ".xlsx")]
                      if data_dir.exists() else [])
    by_stem = {f.stem.lower(): f for f in data_files}
    text = main_md.read_text(encoding="utf-8", errors="replace")
    counters = {"filled": 0}
    unfilled_items = []

    def repl(m):
        spec = m.group(1).strip()
        val = _resolve_data_spec(spec, by_stem)
        if val is None:
            unfilled_items.append(spec)
            return m.group(0)
        counters["filled"] += 1
        return val

    new_text = DATA_PH_RE.sub(repl, text)
    filled = counters["filled"]
    if filled:
        backup = main_md.with_name("main.md.bak")
        if not backup.exists():
            backup.write_text(text, encoding="utf-8")
        main_md.write_text(new_text, encoding="utf-8")
    return {"filled": filled, "unfilled": len(unfilled_items), "items": unfilled_items}


def fill_project(project_dir, values=None):
    project = Path(project_dir)
    data_dir = project / "data"
    if not data_dir.exists():
        return "项目中没有 data/ 目录，请先放入实验数据（.csv/.xlsx）"
    files = [p for p in sorted(data_dir.iterdir()) if p.suffix.lower() in (".csv", ".xlsx")]
    if not files:
        return "data/ 目录为空"
    parts = ["# 自动生成 Results 汇总", ""]
    all_charts = []
    for f in files:
        md = generate_markdown(f, value_cols=values)
        parts.append(md)
        try:
            charts = generate_charts(f, value_cols=values, out_dir=project / "data" / "charts")
            all_charts.extend(charts)
        except Exception as e:
            parts.append(f"\n> 图表生成失败: {e}")
    out_md = project / "manuscript" / "results_auto.md"
    out_md.parent.mkdir(parents=True, exist_ok=True)
    out_md.write_text("\n".join(parts), encoding="utf-8")

    planned = generate_planned_charts(project, files)
    planned_tables = generate_planned_tables(project, files)
    inserted_figs, inserted_tabs = insert_planned_into_manuscript(project, planned, planned_tables)
    kinds = {}
    for t in planned_tables:
        k = t.get("kind", "desc")
        kinds[k] = kinds.get(k, 0) + 1
    table_summary = "表: 描述统计 {} 张 + ANOVA {} 张{}".format(
        kinds.get("desc", 0), kinds.get("anova", 0),
        f"(跳过重复 {kinds.get('dup', 0)} 张)" if kinds.get("dup") else "")
    # 正文级 [DATA:] 占位符回填
    ph = fill_data_placeholders(project, files)
    ph_line = f"[DATA:] 占位符: 已回填 {ph['filled']} 处"
    if ph["unfilled"]:
        ph_line += f"，剩余 {ph['unfilled']} 处待补（缺数据或格式不符）"
    summary = (
        f"已生成: {out_md}\n"
        f"图表: {len(all_charts)} 张（含按 figures.md 规划生成 {len(planned)} 张） -> data/charts/\n"
        f"{table_summary}\n"
        f"已插入正文: 图 {inserted_figs} 张 / 表 {inserted_tabs} 张\n"
        f"{ph_line}"
    )
    return summary


def main():
    ap = argparse.ArgumentParser(prog="data2paper", description="Data-to-Paper 自动统计/图表/Results")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("stats", help="生成统计 Markdown")
    p.add_argument("file")
    p.add_argument("--group", default=None)
    p.add_argument("--values", default=None)
    p.set_defaults(fn=cmd_stats)

    p = sub.add_parser("charts", help="生成图表")
    p.add_argument("file")
    p.add_argument("--group", default=None)
    p.add_argument("--values", default=None)
    p.add_argument("--out", default=None)
    p.set_defaults(fn=cmd_charts)

    p = sub.add_parser("fill", help="扫描项目 data/ 并自动生成 Results")
    p.add_argument("project")
    p.add_argument("--values", default=None)
    p.set_defaults(fn=cmd_fill)

    p = sub.add_parser("fillph", help="仅回填正文 [DATA:] 占位符（不重生成图表）")
    p.add_argument("project")
    p.set_defaults(fn=cmd_fillph)

    p = sub.add_parser("figures", help="按 figures.md 规划混合路由生成图（data/AI 生图/文字图注）")
    p.add_argument("project")
    p.add_argument("--mode", default="auto", choices=["auto", "all"],
                   help="auto=跳过已有产物（默认）; all=强制重生成")
    p.add_argument("--ids", default=None, help="仅处理指定编号，逗号分隔，如: 图1,图3")
    p.add_argument("--json", action="store_true", help="以 JSON 输出完整结果")
    p.set_defaults(fn=cmd_figures)

    args = ap.parse_args()
    args.fn(args)


def _split_values(s):
    if not s:
        return None
    return [x.strip() for x in s.split(",") if x.strip()]


def cmd_stats(args):
    print(generate_markdown(args.file, args.group, _split_values(args.values)))


def cmd_charts(args):
    res = generate_charts(args.file, args.group, _split_values(args.values), args.out)
    for r in res:
        print(r.get("rel") or r.get("file"))


def cmd_fill(args):
    print(fill_project(args.project, _split_values(args.values)))


def cmd_fillph(args):
    r = fill_data_placeholders(args.project)
    if r.get("error"):
        print("错误:", r["error"])
        return
    print(f"[DATA:] 占位符回填: 已填 {r['filled']} 处, 剩余 {r['unfilled']} 处")
    if r["items"]:
        print("未匹配占位符（缺数据或格式不符）:")
        for it in r["items"][:30]:
            print("  -", it)


def cmd_figures(args):
    from web.charts import generate_all_figures
    ids = None
    if args.ids:
        ids = [x.strip() for x in re.split(r"[,，、]", args.ids) if x.strip()]
    res = generate_all_figures(args.project, mode=args.mode, ids=ids)
    if args.json:
        import json as _json
        print(_json.dumps(res, ensure_ascii=False, indent=2))
        return
    print(f"图表混合路由: 共 {res['total']} 张, 生成 {res['generated']}, "
          f"跳过 {res['skipped']}（产物已存在）, 降级 {res['fallback']}")
    for r in res["results"]:
        fig = r.get("fig_id", "?")
        if r.get("skipped"):
            print(f"  - {fig}: 跳过（产物已存在）")
            continue
        route = r.get("route") or "?"
        status = "OK" if r.get("ok") else "失败"
        target = r.get("rel") or r.get("file") or ("文字图注" if r.get("text") else "")
        line = f"  - {fig}: {route} [{status}]"
        if target:
            line += f" -> {target}"
        if r.get("fallback_reason"):
            line += f"（{r['fallback_reason']}）"
        print(line)


if __name__ == "__main__":
    main()
